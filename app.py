# app.py  -  JRC Truck Plan Builder (single Flask app for Vercel)
import io, re, json, base64, tempfile, os
from collections import Counter
from http.server import BaseHTTPRequestHandler
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# ===================== formatter =====================
NEW_COLUMNS = ["JRC EQUIPMENT", "JRC TRUCK", "JRC RATE", "L FT", "W FT", "H FT", "POUNDS"]
ID_HEADER = "ID"

# Header fills: the three JRC-decision columns are green; the four converted-
# measurement columns are yellow. This makes the human-entered/decided section and
# the machine-converted section visually distinct at a glance.
GREEN_HEADERS = {"JRC EQUIPMENT", "JRC TRUCK", "JRC RATE"}
YELLOW_HEADERS = {"L FT", "W FT", "H FT", "POUNDS"}

GREEN = PatternFill(start_color="C1F0C8", end_color="C1F0C8", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
# Over-dimension data cells are BLUE (a load over 12 ft tall, over 8'6" wide, or
# over 47,500 lb); an extreme over-width load (> 16 ft) is RED instead.
BLUE = PatternFill(start_color="83CCEB", end_color="83CCEB", fill_type="solid")
RED = PatternFill(start_color="FF2F2F", end_color="FF2F2F", fill_type="solid")
# Extreme over-width (> 16 ft) cells: YELLOW fill with RED text (replaces the old
# solid-red fill). Header text colours match the header fill (Excel's Good/Neutral
# styles): green headers -> dark-green text, yellow headers -> brown text.
YELLOW_HL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FONT = Font(color="FF0000")
GREEN_FONT = Font(color="006100")
YELLOW_FONT = Font(color="9C5700")

# Practical per-truck cargo payload ceiling (lb). A load at or under this can ride
# a truck; the consolidation grouping must keep each truck's combined cargo weight
# within it, and any single item over it is highlighted blue. This is the federal
# 80,000 lb gross less a typical tractor + open-deck trailer tare.
PAYLOAD_CEILING_LB = 47500

LEN_TO_FT = {
    "mm": 0.00328084, "cm": 0.0328084, "m": 3.2808399, "meter": 3.2808399,
    "metre": 3.2808399, "in": 1/12.0, "inch": 1/12.0, "inches": 1/12.0,
    "\"": 1/12.0, "ft": 1.0, "foot": 1.0, "feet": 1.0, "'": 1.0,
    "yd": 3.0, "yard": 3.0,
}
WT_TO_LB = {
    "g": 0.00220462, "gram": 0.00220462, "kg": 2.20462262, "kilogram": 2.20462262,
    "kgs": 2.20462262, "t": 2204.62262, "tonne": 2204.62262, "tonnes": 2204.62262,
    "metric ton": 2204.62262, "lb": 1.0, "lbs": 1.0, "pound": 1.0, "pounds": 1.0,
    "ton": 2000.0, "us ton": 2000.0, "short ton": 2000.0,
}


def to_ft(meas):
    u = str(meas["unit"]).strip().lower()
    if u not in LEN_TO_FT:
        raise ValueError(f"Unknown length unit {meas['unit']!r}. Flag this; do not guess.")
    return round(float(meas["value"]) * LEN_TO_FT[u], 1)   # one decimal place


def to_lb(meas):
    u = str(meas["unit"]).strip().lower()
    if u not in WT_TO_LB:
        raise ValueError(f"Unknown weight unit {meas['unit']!r}. Flag this; do not guess.")
    return int(round(float(meas["value"]) * WT_TO_LB[u]))   # nearest whole pound


def renumber_by_rule(rows):
    """Number trucks by equipment class: double drops (incl. extended) first,
    then stepdecks, then flatbeds. Within double drops / stepdecks sort by tallest
    load first; within flatbeds sort by longest length first. Ties -> first row."""
    def rank(e):
        e=(e or "").upper()
        if "DOUBLE DROP" in e or "MINIDECK" in e or "MINI DECK" in e: return 0
        if "STEPDECK" in e: return 1
        return 2                            # FLATBED
    groups={}
    for r in rows:
        gid=r["truck"]
        g=groups.setdefault(gid,{"max_h":float("-inf"),"max_l":float("-inf"),
                                 "first_row":r["excel_row"],"equip":r.get("equipment")})
        g["max_h"]=max(g["max_h"],r["h_ft"])
        g["max_l"]=max(g["max_l"],r["l_ft"])
        g["first_row"]=min(g["first_row"],r["excel_row"])
    def sortkey(kv):
        g=kv[1]; rk=rank(g["equip"])
        within = -g["max_l"] if rk==2 else -g["max_h"]   # flatbed by length, else height
        return (rk, within, g["first_row"])
    ordered=sorted(groups.items(), key=sortkey)
    return {gid:i for i,(gid,_) in enumerate(ordered, start=1)}


def renumber_by_height(rows_with_h):
    """Map each model 'truck' id to a final number so the truck carrying the
    tallest load is 1, the next tallest group is 2, etc. Items keep their
    grouping; only the labels change. Ties broken by first appearance (row)."""
    groups = {}
    for r in rows_with_h:
        gid = r["truck"]
        g = groups.setdefault(gid, {"max_h": float("-inf"), "first_row": r["excel_row"]})
        g["max_h"] = max(g["max_h"], r["h_ft"])
        g["first_row"] = min(g["first_row"], r["excel_row"])
    ordered = sorted(groups.items(), key=lambda kv: (-kv[1]["max_h"], kv[1]["first_row"]))
    return {gid: i for i, (gid, _) in enumerate(ordered, start=1)}


def autofit_columns(ws, header_row):
    """Widen every VISIBLE column so the longest cell value is not cut off, and
    wrap the header row. Hidden columns are skipped (left hidden)."""
    from openpyxl.utils import get_column_letter
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].hidden:
            continue
        longest = 0
        for cell in ws[letter]:
            if cell.value is not None:
                for line in str(cell.value).split("\n"):
                    longest = max(longest, len(line))
        ws.column_dimensions[letter].width = min(max(longest + 2, 9), 90)
    # wrap + bold-ish header so long titles stay readable
    for cell in ws[header_row]:
        if cell.value is not None:
            cell.alignment = Alignment(wrap_text=False, vertical="center")
    ws.row_dimensions[header_row].height = 15


def build(spec):
    wb = openpyxl.load_workbook(spec["input_file"])   # keep formatting/structure
    ws = wb[spec["sheet"]]
    header_row = spec.get("header_row", 1)

    original_max_col = ws.max_column

    # 1) Hide every column NOT in keep_visible; make sure kept columns are visible.
    #    We NEVER delete a column. Match by exact header text (model already
    #    resolved fuzzy variants like "case #" -> "Case Number" into actual headers).
    keep = {str(h).strip().lower() for h in spec.get("keep_visible", [])}
    for c in range(1, original_max_col + 1):
        letter = get_column_letter(c)
        header = ws.cell(row=header_row, column=c).value
        header_norm = str(header).strip().lower() if header is not None else ""
        ws.column_dimensions[letter].hidden = header_norm not in keep

    # 2) Append the JRC columns to the RIGHT of all existing columns, in order,
    #    then the identifier column on the far right. Colour the new headers.
    first_new = original_max_col + 1
    headers = NEW_COLUMNS + [ID_HEADER]
    for offset, name in enumerate(headers):
        cell = ws.cell(row=header_row, column=first_new + offset, value=name)
        if name in GREEN_HEADERS:
            cell.fill = GREEN
            cell.font = GREEN_FONT
        elif name in YELLOW_HEADERS:
            cell.fill = YELLOW
            cell.font = YELLOW_FONT
    col_idx = {name: first_new + i for i, name in enumerate(NEW_COLUMNS)}
    id_col = first_new + len(NEW_COLUMNS)

    # 3) Convert + round every row first (we need heights before numbering trucks).
    computed = []
    for row in spec["rows"]:
        computed.append({
            "excel_row": row["excel_row"],
            "l_ft": to_ft(row["length"]),
            "w_ft": to_ft(row["width"]),
            "h_ft": to_ft(row["height"]),
            "lbs": to_lb(row["weight"]),
            "equipment": row.get("equipment"),
            "truck": row.get("truck"),
            "rate": row.get("rate"),
        })

    # 4) Truck numbering: tallest loads get the lowest numbers.
    truck_map = renumber_by_rule(computed)

    # 5) Fill each data row and apply the highlight rules to the converted cells.
    for i, row in enumerate(computed, start=1):
        r = row["excel_row"]
        ws.cell(row=r, column=col_idx["L FT"], value=row["l_ft"])
        ws.cell(row=r, column=col_idx["W FT"], value=row["w_ft"])
        ws.cell(row=r, column=col_idx["H FT"], value=row["h_ft"])
        ws.cell(row=r, column=col_idx["POUNDS"], value=row["lbs"])
        ws.cell(row=r, column=col_idx["JRC EQUIPMENT"], value=row["equipment"])
        ws.cell(row=r, column=col_idx["JRC TRUCK"], value=truck_map[row["truck"]])
        ws.cell(row=r, column=col_idx["JRC RATE"], value=row["rate"])   # blank if None
        ws.cell(row=r, column=id_col, value=i)

        # Height  > 12 ft     -> BLUE the H FT cell.
        # Weight  > 47,500 lb -> BLUE the POUNDS cell (the payload ceiling).
        # Width   > 8.5 ft    -> BLUE the W FT cell; if width > 16 ft, RED instead
        # (red supersedes blue for an extreme over-width load).
        if row["h_ft"] > 12:
            ws.cell(row=r, column=col_idx["H FT"]).fill = BLUE
        if row["lbs"] > PAYLOAD_CEILING_LB:
            ws.cell(row=r, column=col_idx["POUNDS"]).fill = BLUE
        if row["w_ft"] > 16:
            wfc = ws.cell(row=r, column=col_idx["W FT"])
            wfc.fill = YELLOW_HL
            wfc.font = RED_FONT
        elif row["w_ft"] > 8.5:
            ws.cell(row=r, column=col_idx["W FT"]).fill = BLUE

    autofit_columns(ws, header_row)
    wb.save(spec["output_file"])
    return spec["output_file"]



# ===================== 2-D packer =====================
DECKW=8.5
NORMAL={3:32.0,2:29.0,1:48.0,0:53.0}      # normal trailer LENGTH limit per class
def cap_len(cls,ext): return 53.0 if (ext or cls==0) else (48.0 if cls==1 else NORMAL[cls])
def cap_wt(cls): return 35000 if cls in (2,3) else 47500   # double-drop family 35k; flat/step 47.5k

def prep(items):
    # add packing geometry: depth=along-trailer (max horiz dim), lane=across-deck (min horiz dim)
    for r,it in items.items():
        L,W=it['L'],it['W']
        across=min(L,W); depth=max(L,W)
        it['depth']=depth
        it['ow']= across>DECKW+1e-9
        it['lane']= DECKW if it['ow'] else across
        it['owwidth']= across if it['ow'] else 0.0
        it['oversize']= depth>NORMAL[it['rk']]+1e-9   # too long for its band's normal well
    return items

def pack(items, order):
    # EXTENDED trailers (ext double drop / ext minideck) are stretched to fit ONE
    # oversize piece. You may add more pieces ONLY if they do NOT further lengthen the
    # trailer -- i.e. they fit BESIDE the oversize piece within its existing length
    # (an existing shelf), never in a NEW length position. Adding a piece that needs
    # more length is a divisible load and is not permittable.
    # Oversize pieces never join another truck (and two oversize can't share).
    trucks=[]
    for r in order:
        it=items[r]; placed=False
        if not it['oversize']:
            for t in trucks:
                newcls=max(t['rk'],it['rk'])
                if t['wt']+it['wt']>cap_wt(newcls)+1e-6: continue
                if it['ow'] and t['ow_w']>DECKW and it['owwidth']>t['ow_w']+1e-9: continue
                done=False
                if not it['ow']:
                    for sh in t['shelves']:                 # side-by-side within an existing shelf
                        if sh['ow']: continue
                        if sh['depth']+1e-9>=it['depth'] and sh['across']+it['lane']<=DECKW+1e-9:
                            sh['across']+=it['lane']; done=True; break
                if not done and not t['ext']:               # NEW length position only on NON-extended trucks
                    if t['len']+it['depth']<=cap_len(newcls,False)+1e-9:
                        t['shelves'].append({'depth':it['depth'],'across':it['lane'],'ow':it['ow']})
                        t['len']+=it['depth']; done=True
                if done:
                    t['wt']+=it['wt']; t['rk']=newcls
                    if it['ow']: t['ow_w']=max(t['ow_w'],it['owwidth'])
                    t['rows'].append(r); placed=True; break
        if not placed:
            trucks.append({'shelves':[{'depth':it['depth'],'across':it['lane'],'ow':it['ow']}],
                           'len':it['depth'],'wt':it['wt'],'rk':it['rk'],'ext':it['oversize'],
                           'ow_w':it['owwidth'] if it['ow'] else 0.0,'rows':[r]})
    return trucks

def best_pack(items):
    items=prep(items)
    keys={
     'rk,depth':lambda r:(-items[r]['rk'],-items[r]['depth'],-items[r]['wt']),
     'rk,wt':lambda r:(-items[r]['rk'],-items[r]['wt'],-items[r]['depth']),
     'depth':lambda r:(-items[r]['depth'],-items[r]['rk']),
     'rk,ow,depth':lambda r:(-items[r]['rk'],-int(items[r]['ow']),-items[r]['depth']),
    }
    best=None
    for nm,k in keys.items():
        # fresh copy of truck state each run
        for r in items: pass
        t=pack(items,sorted(items,key=k))
        if best is None or len(t)<len(best[1]): best=(nm,t)
    return best

def label(rows,items):
    cls=max(items[x]['rk'] for x in rows); ext=any(items[x]['oversize'] for x in rows)
    if cls==3: return 'EXT MINIDECK' if ext else 'MINIDECK'
    if cls==2: return 'EXTENDED DOUBLE DROP' if ext else 'DOUBLE DROP'
    return 'STEPDECK' if cls==1 else 'FLATBED'

def validate(trucks,items):
    bad=[]
    for t in trucks:
        cls=t['rk']
        if t['wt']>cap_wt(cls)+1 and len(t['rows'])>1: bad.append(('WT',t['rows'],t['wt'],cap_wt(cls)))
        for sh in t['shelves']:
            if sh['across']>DECKW+1e-6: bad.append(('WIDTH',t['rows'],sh['across']))
        if not t['ext'] and t['len']>cap_len(cls,False)+1e-6 and len(t['rows'])>1: bad.append(('LEN',t['rows'],t['len']))
    return bad

if __name__=='__main__':
    import sys
    items={int(k):v for k,v in json.load(open(sys.argv[1])).items()}
    nm,trucks=best_pack(items)
    print('best heuristic',nm,'->',len(trucks),'trucks')
    bad=validate(trucks,items); print('violations',len(bad), bad[:5])
    out=[{'rows':sorted(t['rows']),'eq':label(t['rows'],items)} for t in trucks]
    json.dump(out,open(sys.argv[2],'w'))


# ===================== engine =====================
def load_any(file_bytes, filename):
    """Return (openpyxl_workbook, xlsx_bytes). Handles .xlsx natively and old .xls
    via xlrd (no LibreOffice needed -> works on Vercel serverless)."""
    if filename.lower().endswith(".xls"):
        import xlrd                      # add 'xlrd' to requirements.txt for prod
        book = xlrd.open_workbook(file_contents=file_bytes)
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        for sh in book.sheets():
            ws = wb.create_sheet(sh.name)
            for r in range(sh.nrows):
                for c in range(sh.ncols):
                    ws.cell(r+1, c+1, sh.cell_value(r, c))
        buf = io.BytesIO(); wb.save(buf)
        return openpyxl.load_workbook(io.BytesIO(buf.getvalue()), data_only=True), buf.getvalue()
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True), file_bytes

# ---------------------------------------------------------------- detection ---
DIM_UNIT_RE = re.compile(r"(cm|mm|m\b|in\b|inch|inches|ft|foot|feet)", re.I)

def _norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower()) if s is not None else ""

def find_header_row(ws, max_scan=30):
    """Header row = the first row (within the first few) that looks like column
    titles for length/width/height + a description/case column."""
    best, best_score = 1, -1
    for r in range(1, min(max_scan, ws.max_row) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        norm = [_norm(v) for v in vals]
        score = 0
        for key in ("ln", "length", "lnin", "lncm", "wd", "width", "ht", "height",
                    "description", "casenumber", "grswtlbs", "grosssweight", "weight"):
            if any(key in n for n in norm):
                score += 1
        if score > best_score:
            best, best_score = r, score
    return best

def _match(headers, *cands):
    norm = {i: _norm(h) for i, h in headers.items()}
    # exact-ish contains match, in candidate priority order
    for cand in cands:
        c = _norm(cand)
        for i, n in norm.items():
            if n == c:
                return i
    for cand in cands:
        c = _norm(cand)
        for i, n in norm.items():
            if c and c in n:
                return i
    return None

def detect_columns(ws, header_row):
    headers = {c: ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)
               if ws.cell(header_row, c).value is not None}
    col = {}
    col["desc"] = _match(headers, "description", "box description", "project name", "mli")
    col["case"] = _match(headers, "case number", "box/case number", "order release id",
                          "case no", "case #")
    # prefer imperial dimension columns, fall back to metric
    col["len_in"] = _match(headers, "ln_in", "length (in", "lenght (in")
    col["wd_in"]  = _match(headers, "wd_in", "width (in")
    col["ht_in"]  = _match(headers, "ht_in", "height (in")
    col["wt_lb"]  = _match(headers, "grs_wt_lbs", "gross_weight (lb", "total weight",
                           "gross weight")
    col["len_cm"] = _match(headers, "ln_cm", "lenght (cm", "length (cm")
    col["wd_cm"]  = _match(headers, "wd_cm", "width (cm")
    col["ht_cm"]  = _match(headers, "ht_cm", "height (cm")
    col["wt_kg"]  = _match(headers, "grs_wt_kg", "gross_weight (kg")
    # generic fallbacks
    if col["len_in"] is None and col["len_cm"] is None:
        col["len_in"] = _match(headers, "length", "ln");
    if col["wd_in"] is None and col["wd_cm"] is None:
        col["wd_in"] = _match(headers, "width", "wd")
    if col["ht_in"] is None and col["ht_cm"] is None:
        col["ht_in"] = _match(headers, "height", "ht")
    if col["wt_lb"] is None and col["wt_kg"] is None:
        col["wt_lb"] = _match(headers, "weight")
    return headers, col

def pick_dims(col):
    """Return (len_col, wd_col, ht_col, wt_col, unit_len, unit_wt) preferring imperial."""
    if col["len_in"] and col["wd_in"] and col["ht_in"]:
        ul = "in"
    elif col["len_cm"] and col["wd_cm"] and col["ht_cm"]:
        return col["len_cm"], col["wd_cm"], col["ht_cm"], (col["wt_kg"] or col["wt_lb"]), "cm", ("kg" if col["wt_kg"] else "lb")
    else:
        ul = "in"
    return col["len_in"], col["wd_in"], col["ht_in"], (col["wt_lb"] or col["wt_kg"]), ul, ("lb" if col["wt_lb"] else "kg")

# ----------------------------------------------------------------- numbers ---
def num(v):
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return None

def feet(v, unit):
    f = {"in": 1/12.0, "cm": 0.0328084, "mm": 0.00328084, "m": 3.2808399, "ft": 1.0}[unit]
    return v * f

# --------------------------------------------------------------- equipment ---
def bclass(h_ft):
    return 3 if h_ft > 14.5 else 2 if h_ft > 10 else 1 if h_ft > 8.5 else 0

# -------------------------------------------------------------- main entry ---
def _scan(ws):
    hr = find_header_row(ws)
    headers, col = detect_columns(ws, hr)
    lc, wc, hc, wtc, ul, uw = pick_dims(col)
    return hr, headers, col, lc, wc, hc, wtc, ul, uw

def _rows_for(ws, hr, col, lc, wc, hc, wtc, ul, uw):
    items, skip = {}, {"consolidated": 0, "zero": 0}
    for r in range(hr + 1, ws.max_row + 1):
        lv = ws.cell(r, lc).value
        case = ws.cell(r, col["case"]).value if col["case"] else None
        desc = ws.cell(r, col["desc"]).value if col["desc"] else None
        if lv is None and case is None and desc is None:
            continue
        if case and "onsolidat" in str(case):
            skip["consolidated"] += 1; continue
        L, W, H, WT = num(lv), num(ws.cell(r, wc).value), num(ws.cell(r, hc).value), num(ws.cell(r, wtc).value)
        if (L in (0, None)) and (WT in (0, None)):
            skip["zero"] += 1; continue
        Lf, Wf, Hf = feet(L, ul), feet(W, ul), feet(H, ul)
        wt_lb = WT if uw == "lb" else WT * 2.20462262
        items[r] = dict(L=Lf, W=Wf, H=Hf, wt=wt_lb, rk=bclass(Hf), ow=Wf > 8.5)
    return items, skip

def process(file_bytes, filename="manifest.xlsx"):
    """Run the full pipeline over EVERY cargo tab. Returns (xlsx_bytes, summary)."""
    import tempfile, os
    from collections import Counter
    wb, xlsx_bytes = load_any(file_bytes, filename)

    cargo = []
    for sh in wb.worksheets:
        if sh.max_row < 2:
            continue
        hr, headers, col, lc, wc, hc, wtc, ul, uw = _scan(sh)
        if all([lc, wc, hc, wtc]):
            cargo.append((sh, hr, headers, col, lc, wc, hc, wtc, ul, uw))
    if not cargo:
        raise ValueError("Couldn't find a tab with Length/Width/Height/Weight columns. "
                         "If the cargo list is on a specific tab, make sure that tab has those headers.")

    cur = xlsx_bytes
    tot_pieces = tot_trucks = tot_ow = 0; tot_wt = 0.0
    mix = Counter(); superloads = []; skipped = {"consolidated": 0, "zero": 0}; tabs = []
    for (sh, hr, headers, col, lc, wc, hc, wtc, ul, uw) in cargo:
        items, skip = _rows_for(sh, hr, col, lc, wc, hc, wtc, ul, uw)
        if not items:
            continue
        _, trucks = best_pack(items)
        groups = [{"rows": sorted(t["rows"]), "eq": label(t["rows"], items)} for t in trucks]
        keep = []
        for _k, idx in [("desc", col["desc"]), ("case", col["case"]),
                        (None, lc), (None, wc), (None, hc), (None, wtc)]:
            if idx:
                keep.append(headers[idx])
        rows_spec, gid = [], 0
        for g in groups:
            gid += 1
            for r in g["rows"]:
                rows_spec.append({"excel_row": r,
                    "length": {"value": num(sh.cell(r, lc).value), "unit": ul},
                    "width":  {"value": num(sh.cell(r, wc).value), "unit": ul},
                    "height": {"value": num(sh.cell(r, hc).value), "unit": ul},
                    "weight": {"value": num(sh.cell(r, wtc).value), "unit": uw},
                    "equipment": g["eq"], "truck": "g%03d" % gid, "rate": None})
        rows_spec.sort(key=lambda x: x["excel_row"])
        tin = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); tin.write(cur); tin.close()
        tout = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); tout.close()
        build({"input_file": tin.name, "output_file": tout.name, "sheet": sh.title,
               "header_row": hr, "keep_visible": keep, "rows": rows_spec})
        cur = open(tout.name, "rb").read(); os.unlink(tin.name); os.unlink(tout.name)

        s = make_summary(items, groups, skip, len(items))
        tot_pieces += s["pieces"]; tot_trucks += s["trucks"]; tot_ow += s["over_width_pieces"]; tot_wt += s["total_weight_lb"]
        for k, v in s["equipment_mix"].items():
            mix[k] += v
        for row, reason in s["superloads"]:
            superloads.append(("%s (%s)" % (row, sh.title), reason))
        skipped["consolidated"] += skip["consolidated"]; skipped["zero"] += skip["zero"]
        tabs.append({"tab": sh.title, "trucks": s["trucks"], "pieces": s["pieces"]})

    if tot_pieces == 0:
        raise ValueError("No cargo rows found.")
    summary = {"pieces": tot_pieces, "trucks": tot_trucks, "equipment_mix": dict(mix),
               "over_width_pieces": tot_ow, "superloads": superloads, "skipped": skipped,
               "total_weight_lb": round(tot_wt), "tabs": tabs}
    return cur, summary

# ----------------------------------------------------------------- summary ---
def make_summary(items, groups, skipped, n):
    from collections import Counter
    mix = Counter(g["eq"] for g in groups)
    ow = sum(1 for i in items.values() if i["W"] > 8.5)
    superload = []
    for r, i in items.items():
        if i["H"] > 13.5: superload.append((r, "over-height %.1f ft" % i["H"]))
        if i["W"] > 16: superload.append((r, "over-width %.1f ft" % i["W"]))
        if i["wt"] > 47500: superload.append((r, "heavy-haul %.0f lb" % i["wt"]))
        if max(i["L"], i["W"]) > 55: superload.append((r, "over-length %.1f ft" % max(i["L"], i["W"])))
    return {
        "pieces": n,
        "trucks": len(groups),
        "equipment_mix": dict(mix),
        "over_width_pieces": ow,
        "superloads": superload,
        "skipped": skipped,
        "total_weight_lb": round(sum(i["wt"] for i in items.values())),
    }

# ===================== Flask app (single entrypoint: app.py) =====================
import os, base64 as _b64
from flask import Flask, request, jsonify, send_file, Response
app = Flask(__name__)

_PAGE = _b64.b64decode("PCFkb2N0eXBlIGh0bWw+CjxodG1sIGxhbmc9ImVuIj4KPGhlYWQ+CjxtZXRhIGNoYXJzZXQ9InV0Zi04Ij4KPG1ldGEgbmFtZT0idmlld3BvcnQiIGNvbnRlbnQ9IndpZHRoPWRldmljZS13aWR0aCxpbml0aWFsLXNjYWxlPTEiPgo8dGl0bGU+VHJ1Y2sgUGxhbiBCdWlsZGVyICBKUkMgVHJhbnNwb3J0YXRpb248L3RpdGxlPgo8bGluayByZWw9InByZWNvbm5lY3QiIGhyZWY9Imh0dHBzOi8vZm9udHMuZ29vZ2xlYXBpcy5jb20iPgo8bGluayByZWw9InByZWNvbm5lY3QiIGhyZWY9Imh0dHBzOi8vZm9udHMuZ3N0YXRpYy5jb20iIGNyb3Nzb3JpZ2luPgo8bGluayBocmVmPSJodHRwczovL2ZvbnRzLmdvb2dsZWFwaXMuY29tL2NzczI/ZmFtaWx5PUludGVyOndnaHRANDAwOzUwMDs2MDA7NzAwJmRpc3BsYXk9c3dhcCIgcmVsPSJzdHlsZXNoZWV0Ij4KPHN0eWxlPgogOnJvb3R7CiAgIC0tYmc6IzBiMTIyMDsgLS1iZzI6IzBmMWEzMDsgLS1wYW5lbDojMGYxODMwOyAtLXBhbmVsMjojMTMyMDNjOwogICAtLWxpbmU6cmdiYSgyNTUsMjU1LDI1NSwuMDkpOyAtLWxpbmUyOnJnYmEoMjU1LDI1NSwyNTUsLjE2KTsKICAgLS1pbms6I2VlZjJmOTsgLS1zdWI6IzlhYTdjMjsgLS1oaW50OiM2YzdhOTk7CiAgIC0tYnJhbmQ6IzNiODJmNjsgLS1icmFuZDI6IzYwYTVmYTsgLS1icmFuZGluazojYmNkNGZmOwogICAtLW9rOiMzNGQzOTk7IC0tb2tiZzpyZ2JhKDUyLDIxMSwxNTMsLjEyKTsgLS13YXJuOiNmODcxNzE7IC0td2FybmJnOnJnYmEoMjQ4LDExMywxMTMsLjEyKTsKICAgLS1yOjE0cHg7IC0tcjI6MjBweDsKIH0KICp7Ym94LXNpemluZzpib3JkZXItYm94fQogaHRtbCxib2R5e21hcmdpbjowO2hlaWdodDoxMDAlfQogYm9keXsKICAgZm9udC1mYW1pbHk6SW50ZXIsLWFwcGxlLXN5c3RlbSxCbGlua01hY1N5c3RlbUZvbnQsIlNlZ29lIFVJIixSb2JvdG8sc2Fucy1zZXJpZjsKICAgY29sb3I6dmFyKC0taW5rKTtiYWNrZ3JvdW5kOnZhcigtLWJnKTsKICAgYmFja2dyb3VuZC1pbWFnZTpyYWRpYWwtZ3JhZGllbnQoOTAwcHggNTAwcHggYXQgODAlIC0xMCUscmdiYSg1OSwxMzAsMjQ2LC4xOCksdHJhbnNwYXJlbnQgNjAlKSwKICAgICAgICAgICAgICAgICAgICByYWRpYWwtZ3JhZGllbnQoNzAwcHggNTAwcHggYXQgMCUgMCUscmdiYSg5OSwxMDIsMjQxLC4xMiksdHJhbnNwYXJlbnQgNTUlKTsKICAgLXdlYmtpdC1mb250LXNtb290aGluZzphbnRpYWxpYXNlZDtsaW5lLWhlaWdodDoxLjU7CiB9CiAud3JhcHttYXgtd2lkdGg6MTEwMHB4O21hcmdpbjowIGF1dG87cGFkZGluZzozMHB4IDM2cHggNzBweH0KIG5hdntkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDoxMXB4O2NvbG9yOiNmZmZ9CiAubG9nb3t3aWR0aDozOHB4O2hlaWdodDozOHB4O2JvcmRlci1yYWRpdXM6MTFweDtiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxNDBkZWcsIzI1NjNlYiwjN2MzYWVkKTsKICAgZGlzcGxheTpncmlkO3BsYWNlLWl0ZW1zOmNlbnRlcjtmb250LXNpemU6MjBweDtib3gtc2hhZG93OjAgOHB4IDI0cHggLThweCAjMjU2M2ViODh9CiBuYXYgLm5hbWV7Zm9udC13ZWlnaHQ6NzAwO2ZvbnQtc2l6ZToxNnB4O2xldHRlci1zcGFjaW5nOi4ycHh9CiBuYXYgLnRhZ3tmb250LXNpemU6MTJweDtjb2xvcjp2YXIoLS1zdWIpfQogbmF2IC5waWxse21hcmdpbi1sZWZ0OmF1dG87Zm9udC1zaXplOjEycHg7Y29sb3I6dmFyKC0tYnJhbmRpbmspO2JhY2tncm91bmQ6cmdiYSg1OSwxMzAsMjQ2LC4xNCk7CiAgIGJvcmRlcjoxcHggc29saWQgcmdiYSg1OSwxMzAsMjQ2LC4zKTtib3JkZXItcmFkaXVzOjMwcHg7cGFkZGluZzo1cHggMTJweH0KIC5oZXJve3RleHQtYWxpZ246bGVmdDttYXJnaW46MH0KIC5oZXJvIGgxe2ZvbnQtc2l6ZTozMHB4O2xpbmUtaGVpZ2h0OjEuMTttYXJnaW46MCAwIDEwcHg7bGV0dGVyLXNwYWNpbmc6LS42cHg7Zm9udC13ZWlnaHQ6NzAwO2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEyMGRlZywjZmZmLCNiY2Q0ZmYpOy13ZWJraXQtYmFja2dyb3VuZC1jbGlwOnRleHQ7YmFja2dyb3VuZC1jbGlwOnRleHQ7Y29sb3I6dHJhbnNwYXJlbnR9CiAuaGVybyBwe21hcmdpbjowO21heC13aWR0aDo1MjBweDtjb2xvcjp2YXIoLS1zdWIpO2ZvbnQtc2l6ZToxNnB4O2xpbmUtaGVpZ2h0OjEuNTV9CiAucGFuZWx7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTgwZGVnLHZhcigtLXBhbmVsKSx2YXIoLS1wYW5lbDIpKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWxpbmUpOwogICBib3JkZXItcmFkaXVzOnZhcigtLXIyKTtwYWRkaW5nOjEwcHg7Ym94LXNoYWRvdzowIDMwcHggODBweCAtNDBweCAjMDAwYSwgaW5zZXQgMCAxcHggMCByZ2JhKDI1NSwyNTUsMjU1LC4wNCl9CiAuZHJvcHtib3JkZXI6MS41cHggZGFzaGVkIHZhcigtLWxpbmUyKTtib3JkZXItcmFkaXVzOnZhcigtLXIpO3BhZGRpbmc6NjRweCAyMnB4O3RleHQtYWxpZ246Y2VudGVyO2N1cnNvcjpwb2ludGVyO3RyYW5zaXRpb246LjE4cztiYWNrZ3JvdW5kOnJnYmEoMjU1LDI1NSwyNTUsLjAxNSl9CiAuZHJvcDpob3Zlcntib3JkZXItY29sb3I6dmFyKC0tYnJhbmQyKTtiYWNrZ3JvdW5kOnJnYmEoNTksMTMwLDI0NiwuMDYpfQogLmRyb3Aub3Zlcntib3JkZXItY29sb3I6dmFyKC0tYnJhbmQyKTtiYWNrZ3JvdW5kOnJnYmEoNTksMTMwLDI0NiwuMSk7dHJhbnNmb3JtOnNjYWxlKDEuMDA0KX0KIC5kcm9wIC5pY3t3aWR0aDo1OHB4O2hlaWdodDo1OHB4O2JvcmRlci1yYWRpdXM6MTZweDttYXJnaW46MCBhdXRvIDE0cHg7ZGlzcGxheTpncmlkO3BsYWNlLWl0ZW1zOmNlbnRlcjsKICAgZm9udC1zaXplOjI3cHg7YmFja2dyb3VuZDpyZ2JhKDU5LDEzMCwyNDYsLjE0KTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoNTksMTMwLDI0NiwuMjgpO2NvbG9yOnZhcigtLWJyYW5kMil9CiAuZHJvcCAudHtmb250LXNpemU6MTdweDtmb250LXdlaWdodDo2MDB9IC5kcm9wIC5ze2NvbG9yOnZhcigtLXN1Yik7Zm9udC1zaXplOjEzcHg7bWFyZ2luLXRvcDo1cHh9CiAuYWN0aW9uc3tkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDoxMnB4O3BhZGRpbmc6MTRweCAxMnB4IDZweDtmbGV4LXdyYXA6d3JhcH0KIGJ1dHRvbntmb250LWZhbWlseTppbmhlcml0O2ZvbnQtd2VpZ2h0OjYwMDtmb250LXNpemU6MTVweDtib3JkZXI6MDtib3JkZXItcmFkaXVzOjExcHg7cGFkZGluZzoxM3B4IDIycHg7Y3Vyc29yOnBvaW50ZXI7dHJhbnNpdGlvbjouMTZzfQogLnByaW1hcnl7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTgwZGVnLCMzYjgyZjYsIzI1NjNlYik7Y29sb3I6I2ZmZjtib3gtc2hhZG93OjAgMTJweCAyNnB4IC0xMnB4ICMyNTYzZWJjY30KIC5wcmltYXJ5OmhvdmVyOm5vdCg6ZGlzYWJsZWQpe2ZpbHRlcjpicmlnaHRuZXNzKDEuMDcpfSBidXR0b246ZGlzYWJsZWR7b3BhY2l0eTouNDtjdXJzb3I6ZGVmYXVsdDtib3gtc2hhZG93Om5vbmV9CiAuZ2hvc3R7YmFja2dyb3VuZDpyZ2JhKDI1NSwyNTUsMjU1LC4wNSk7Y29sb3I6dmFyKC0taW5rKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWxpbmUpfQogLnN0YXR1c3tkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo5cHg7Y29sb3I6dmFyKC0tc3ViKTtmb250LXNpemU6MTRweH0KIC5zcGlue3dpZHRoOjE2cHg7aGVpZ2h0OjE2cHg7Ym9yZGVyOjJweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4xOCk7Ym9yZGVyLXRvcC1jb2xvcjp2YXIoLS1icmFuZDIpO2JvcmRlci1yYWRpdXM6NTAlO2FuaW1hdGlvbjpzcCAuN3MgbGluZWFyIGluZmluaXRlfQogQGtleWZyYW1lcyBzcHt0b3t0cmFuc2Zvcm06cm90YXRlKDM2MGRlZyl9fQogLmZlYXRze2Rpc3BsYXk6Z3JpZDtncmlkLXRlbXBsYXRlLWNvbHVtbnM6cmVwZWF0KDMsMWZyKTtnYXA6MTRweDttYXJnaW46MjZweCAwIDB9CiAuZmVhdHtiYWNrZ3JvdW5kOnJnYmEoMjU1LDI1NSwyNTUsLjAyNSk7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1saW5lKTtib3JkZXItcmFkaXVzOnZhcigtLXIpO3BhZGRpbmc6MTVweCAxNHB4fQogLmZlYXQgLmZpe2NvbG9yOnZhcigtLWJyYW5kMik7Zm9udC1zaXplOjE4cHh9IC5mZWF0IGg0e21hcmdpbjo4cHggMCAzcHg7Zm9udC1zaXplOjE0cHg7Zm9udC13ZWlnaHQ6NjAwfQogLmZlYXQgcHttYXJnaW46MDtjb2xvcjp2YXIoLS1zdWIpO2ZvbnQtc2l6ZToxMi41cHg7bGluZS1oZWlnaHQ6MS41fQogLnJlc3VsdHttYXJnaW4tdG9wOjE4cHh9CiAucmNhcmR7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTgwZGVnLHZhcigtLXBhbmVsKSx2YXIoLS1wYW5lbDIpKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWxpbmUpO2JvcmRlci1yYWRpdXM6dmFyKC0tcjIpO3BhZGRpbmc6MjJweH0KIC5yaGVhZHtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDoxMHB4O2ZvbnQtd2VpZ2h0OjYwMDtmb250LXNpemU6MTZweDttYXJnaW4tYm90dG9tOjE4cHh9CiAuY2hlY2t7d2lkdGg6MjZweDtoZWlnaHQ6MjZweDtib3JkZXItcmFkaXVzOjUwJTtiYWNrZ3JvdW5kOnZhcigtLW9rYmcpO2NvbG9yOnZhcigtLW9rKTtkaXNwbGF5OmdyaWQ7cGxhY2UtaXRlbXM6Y2VudGVyO2ZvbnQtc2l6ZToxNXB4fQogLnN0YXRze2Rpc3BsYXk6Z3JpZDtncmlkLXRlbXBsYXRlLWNvbHVtbnM6cmVwZWF0KDQsMWZyKTtnYXA6MTJweH0KIC5zdGF0e2JhY2tncm91bmQ6cmdiYSgyNTUsMjU1LDI1NSwuMDMpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tbGluZSk7Ym9yZGVyLXJhZGl1czp2YXIoLS1yKTtwYWRkaW5nOjE2cHggMTRweH0KIC5zdGF0IC5ue2ZvbnQtc2l6ZTozMHB4O2ZvbnQtd2VpZ2h0OjcwMDtsZXR0ZXItc3BhY2luZzotMXB4fSAuc3RhdCAubHtjb2xvcjp2YXIoLS1zdWIpO2ZvbnQtc2l6ZToxMS41cHg7dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO2xldHRlci1zcGFjaW5nOi42cHg7bWFyZ2luLXRvcDozcHh9CiAubWl4e2Rpc3BsYXk6ZmxleDtmbGV4LXdyYXA6d3JhcDtnYXA6OHB4O21hcmdpbjoxOHB4IDB9CiAucGlsbDJ7ZGlzcGxheTppbmxpbmUtZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjdweDtiYWNrZ3JvdW5kOnJnYmEoMjU1LDI1NSwyNTUsLjA0KTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWxpbmUpO2JvcmRlci1yYWRpdXM6MzBweDtwYWRkaW5nOjdweCAxM3B4O2ZvbnQtc2l6ZToxM3B4fQogLnBpbGwyIGJ7YmFja2dyb3VuZDp2YXIoLS1icmFuZCk7Y29sb3I6I2ZmZjtib3JkZXItcmFkaXVzOjIwcHg7bWluLXdpZHRoOjIwcHg7dGV4dC1hbGlnbjpjZW50ZXI7cGFkZGluZzoxcHggN3B4O2ZvbnQtc2l6ZToxMnB4fQogLmJhbm5lcntib3JkZXItcmFkaXVzOnZhcigtLXIpO3BhZGRpbmc6MTRweCAxNnB4O21hcmdpbjo0cHggMCAxOHB4O2ZvbnQtc2l6ZToxNHB4O2Rpc3BsYXk6ZmxleDtnYXA6MTFweH0KIC5iYW5uZXIud2FybntiYWNrZ3JvdW5kOnZhcigtLXdhcm5iZyk7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDI0OCwxMTMsMTEzLC4zKTtjb2xvcjojZmVjYWNhfQogLmJhbm5lci5va3tiYWNrZ3JvdW5kOnZhcigtLW9rYmcpO2JvcmRlcjoxcHggc29saWQgcmdiYSg1MiwyMTEsMTUzLC4zKTtjb2xvcjojYmJmN2QwfQogLmJhbm5lciBpe2ZvbnQtc2l6ZToxOHB4O21hcmdpbi10b3A6MXB4fQogLmJhbm5lciAudHRse2ZvbnQtd2VpZ2h0OjYwMH0KIC5kbHtkaXNwbGF5OmlubGluZS1mbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6OXB4O2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDE4MGRlZywjMTBiOTgxLCMwNTk2NjkpO2NvbG9yOiNmZmY7CiAgIGJvcmRlci1yYWRpdXM6MTFweDtwYWRkaW5nOjEzcHggMjJweDtmb250LXdlaWdodDo2MDA7dGV4dC1kZWNvcmF0aW9uOm5vbmU7Zm9udC1zaXplOjE1cHg7Ym94LXNoYWRvdzowIDEycHggMjZweCAtMTJweCAjMDU5NjY5Y2N9CiAuZGw6aG92ZXJ7ZmlsdGVyOmJyaWdodG5lc3MoMS4wNyl9CiAuZXJye2JhY2tncm91bmQ6dmFyKC0td2FybmJnKTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjQ4LDExMywxMTMsLjMpO2NvbG9yOiNmZWNhY2E7Ym9yZGVyLXJhZGl1czp2YXIoLS1yKTtwYWRkaW5nOjE1cHh9CiAuZm9vdHt0ZXh0LWFsaWduOmNlbnRlcjtjb2xvcjp2YXIoLS1oaW50KTtmb250LXNpemU6MTJweDttYXJnaW4tdG9wOjM0cHh9CiAuaGlkZXtkaXNwbGF5Om5vbmV9CiBAbWVkaWEobWF4LXdpZHRoOjU2MHB4KXsuc3RhdHN7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOnJlcGVhdCgyLDFmcil9LmZlYXRze2dyaWQtdGVtcGxhdGUtY29sdW1uczoxZnJ9Lmhlcm8gaDF7Zm9udC1zaXplOjMycHh9fQogLndlbGNvbWV7ZGlzcGxheTpmbGV4O2ZsZXgtZGlyZWN0aW9uOmNvbHVtbjthbGlnbi1pdGVtczpmbGV4LXN0YXJ0O2dhcDoxMHB4O3BhZGRpbmc6MCAwIDE4cHh9CiAud2VsY29tZSBpbWd7ZGlzcGxheTpibG9jaztoZWlnaHQ6YXV0bzt3aWR0aDoxMDAlO21heC13aWR0aDo0NDBweDtpbWFnZS1yZW5kZXJpbmc6YXV0b30KIC53ZWxjb21lIC53dHh0e2ZvbnQtc2l6ZTo0NnB4O2ZvbnQtd2VpZ2h0OjcwMDtsZXR0ZXItc3BhY2luZzotMXB4O2xpbmUtaGVpZ2h0OjE7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTIwZGVnLCNmZmYsI2JjZDRmZik7LXdlYmtpdC1iYWNrZ3JvdW5kLWNsaXA6dGV4dDtiYWNrZ3JvdW5kLWNsaXA6dGV4dDtjb2xvcjp0cmFuc3BhcmVudH0KIC5sZWFke2Rpc3BsYXk6Z3JpZDtncmlkLXRlbXBsYXRlLWNvbHVtbnM6MS4wNWZyIC45NWZyO2dhcDozNHB4O2FsaWduLWl0ZW1zOmNlbnRlcjttYXJnaW4tdG9wOjZweH0gLmxlYWQgLmxlZnR7bWluLXdpZHRoOjB9IC5sZWFkIC5yaWdodHttaW4td2lkdGg6MH0gLnBhbmVse21hcmdpbjowfSBAbWVkaWEobWF4LXdpZHRoOjg2MHB4KXsud3JhcHtwYWRkaW5nOjI0cHggMThweCA2MHB4fS5sZWFke2dyaWQtdGVtcGxhdGUtY29sdW1uczoxZnI7Z2FwOjIycHg7YWxpZ24taXRlbXM6c3RyZXRjaH0uaGVyb3t0ZXh0LWFsaWduOmNlbnRlcn0ud2VsY29tZXthbGlnbi1pdGVtczpjZW50ZXJ9LndlbGNvbWUgaW1ne21heC13aWR0aDo4NnZ3fS5oZXJvIHB7bWFyZ2luOjAgYXV0b30uZmVhdHN7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOjFmcjttYXJnaW4tdG9wOjE2cHh9fS5sZWFke2dyaWQtdGVtcGxhdGUtY29sdW1uczoxZnI7Z2FwOjIycHh9Lmhlcm97dGV4dC1hbGlnbjpjZW50ZXJ9LndlbGNvbWV7YWxpZ24taXRlbXM6Y2VudGVyfS53ZWxjb21lIGltZ3ttYXgtd2lkdGg6ODZ2d30uaGVybyBwe21hcmdpbjowIGF1dG99LmxlYWQgLnJpZ2h0e3Bvc2l0aW9uOnN0YXRpY30uZmVhdHN7bWFyZ2luLXRvcDoxNHB4fX0KPC9zdHlsZT4KPC9oZWFkPgo8Ym9keT4KPGRpdiBjbGFzcz0id3JhcCI+CiAKIAogPG5hdj4KICAgPGRpdiBjbGFzcz0ibG9nbyI+JiMxMjg2NjY7PC9kaXY+CiAgIDxkaXY+PGRpdiBjbGFzcz0ibmFtZSI+SlJDJm5ic3A7VHJhbnNwb3J0YXRpb248L2Rpdj48ZGl2IGNsYXNzPSJ0YWciPlRydWNrIFBsYW4gQnVpbGRlcjwvZGl2PjwvZGl2PgogICA8ZGl2IGNsYXNzPSJwaWxsIj5CZXRhPC9kaXY+CiA8L25hdj4KIDxkaXYgY2xhc3M9ImxlYWQiPjxkaXYgY2xhc3M9ImxlZnQiPjxkaXYgY2xhc3M9IndlbGNvbWUiPjxzcGFuIGNsYXNzPSJ3dHh0Ij5XZWxjb21lIERSUiE8L3NwYW4+PGltZyBzcmM9IndlbGNvbWUucG5nIiBhbHQ9IkRSUiB3ZWxjb21lIiBvbmVycm9yPSJ0aGlzLnN0eWxlLmRpc3BsYXk9J25vbmUnIj48L2Rpdj48ZGl2IGNsYXNzPSJoZXJvIj4KICAgPGgxPkZyb20gbWFuaWZlc3QgdG8gdHJ1Y2sgcGxhbiBpbiBzZWNvbmRzPC9oMT4KICAgPHA+VXBsb2FkIGEgdmVzc2VsIG1hbmlmZXN0IG9yIHBhY2tpbmcgbGlzdC4gV2UgYXNzaWduIGV2ZXJ5IHBpZWNlIHRvIHRoZSByaWdodCB0cmFpbGVyLCBjb25zb2xpZGF0ZSBvbnRvIHRoZSBmZXdlc3QgbGVnYWwgdHJ1Y2tzLCBhbmQgZmxhZyBwZXJtaXQgbG9hZHMgYmVmb3JlIGhhbmRpbmcgYmFjayBhIGZpbmlzaGVkIHdvcmtib29rLjwvcD4KIDwvZGl2PjwvZGl2PjxkaXYgY2xhc3M9InJpZ2h0Ij48ZGl2IGNsYXNzPSJwYW5lbCI+CiAgIDxkaXYgaWQ9ImRyb3AiIGNsYXNzPSJkcm9wIj4KICAgICA8ZGl2IGNsYXNzPSJpYyI+JiMxMTAxNDs8L2Rpdj4KICAgICA8ZGl2IGNsYXNzPSJ0IiBpZD0iZHJvcHQiPkRyb3AgeW91ciBzcHJlYWRzaGVldCBoZXJlPC9kaXY+CiAgICAgPGRpdiBjbGFzcz0icyIgaWQ9ImRyb3BzIj5vciBjbGljayB0byBicm93c2UgJm5ic3A7Jm1pZGRvdDsmbmJzcDsgLnhsc3ggb3IgLnhsczwvZGl2PgogICA8L2Rpdj4KICAgPGlucHV0IGlkPSJmaWxlIiB0eXBlPSJmaWxlIiBhY2NlcHQ9Ii54bHN4LC54bHMiIGNsYXNzPSJoaWRlIj4KICAgPGRpdiBjbGFzcz0iYWN0aW9ucyI+CiAgICAgPGJ1dHRvbiBjbGFzcz0icHJpbWFyeSIgaWQ9ImdvIiBkaXNhYmxlZD5CdWlsZCB0cnVjayBwbGFuPC9idXR0b24+CiAgICAgPGJ1dHRvbiBjbGFzcz0iZ2hvc3QgaGlkZSIgaWQ9InJlc2V0Ij5TdGFydCBvdmVyPC9idXR0b24+CiAgICAgPHNwYW4gY2xhc3M9InN0YXR1cyIgaWQ9InN0YXR1cyI+PC9zcGFuPgogICA8L2Rpdj4KIDwvZGl2PjwvZGl2PjwvZGl2PgogCgogCgogCgogCiAgIDxkaXYgY2xhc3M9ImZlYXQiPjxkaXYgY2xhc3M9ImZpIj4mIzk4ODg7PC9kaXY+PGg0PlBlcm1pdCBmbGFnczwvaDQ+PHA+T3Zlci13aWR0aCwgb3Zlci1oZWlnaHQgYW5kIGhlYXZ5LWhhdWwgc3VwZXJsb2FkcyBjYWxsZWQgb3V0IGF1dG9tYXRpY2FsbHkuPC9wPjwvZGl2PgogPC9kaXY+CgogPGRpdiBjbGFzcz0iZmVhdHMiIGlkPSJmZWF0cyI+CiAgIDxkaXYgY2xhc3M9ImZlYXQiPjxkaXYgY2xhc3M9ImZpIj4mIzk4ODE7PC9kaXY+PGg0PkF1dG8gdHJhaWxlciBzZWxlY3Rpb248L2g0PjxwPkZsYXRiZWQsIHN0ZXBkZWNrLCBkb3VibGUgZHJvcCwgbWluaWRlY2sgIGNob3NlbiBieSBoZWlnaHQsIGxlbmd0aCBhbmQgd2lkdGguPC9wPjwvZGl2PgogICA8ZGl2IGNsYXNzPSJmZWF0Ij48ZGl2IGNsYXNzPSJmaSI+JiMxMjgyMzA7PC9kaXY+PGg0PjItRCBjb25zb2xpZGF0aW9uPC9oND48cD5TaWRlLWJ5LXNpZGUgYW5kIGVuZC10by1lbmQgcGFja2luZyB3aXRoaW4gZXZlcnkgbGVnYWwgbGVuZ3RoLCB3aWR0aCBhbmQgd2VpZ2h0IGxpbWl0LjwvcD48L2Rpdj4KIDxkaXYgaWQ9InJlc3VsdCIgY2xhc3M9InJlc3VsdCI+PC9kaXY+CiA8ZGl2IGNsYXNzPSJmb290Ij5KUkMgVHJhbnNwb3J0YXRpb24gJm1pZGRvdDsgdHJhaWxlciBzZWxlY3Rpb24gJmFtcDsgY29uc29saWRhdGlvbiBlbmdpbmU8L2Rpdj4KPC9kaXY+Cgo8c2NyaXB0Pgpjb25zdCAkPXM9PmRvY3VtZW50LnF1ZXJ5U2VsZWN0b3Iocyk7CmNvbnN0IGRyb3A9JCgiI2Ryb3AiKSxmaWxlPSQoIiNmaWxlIiksZ289JCgiI2dvIikscmVzZXQ9JCgiI3Jlc2V0Iiksc3RhdHVzPSQoIiNzdGF0dXMiKSxyZXN1bHQ9JCgiI3Jlc3VsdCIpLGZlYXRzPSQoIiNmZWF0cyIpOwpsZXQgY2hvc2VuPW51bGw7CmRyb3Aub25jbGljaz0oKT0+ZmlsZS5jbGljaygpOwpbImRyYWdvdmVyIiwiZHJhZ2xlYXZlIiwiZHJvcCJdLmZvckVhY2goZT0+ZHJvcC5hZGRFdmVudExpc3RlbmVyKGUsZXY9Pntldi5wcmV2ZW50RGVmYXVsdCgpO2Ryb3AuY2xhc3NMaXN0LnRvZ2dsZSgib3ZlciIsZT09PSJkcmFnb3ZlciIpO30pKTsKZHJvcC5hZGRFdmVudExpc3RlbmVyKCJkcm9wIixldj0+e2lmKGV2LmRhdGFUcmFuc2Zlci5maWxlc1swXSl7Y2hvc2VuPWV2LmRhdGFUcmFuc2Zlci5maWxlc1swXTtwaWNrKCk7fX0pOwpmaWxlLm9uY2hhbmdlPSgpPT57aWYoZmlsZS5maWxlc1swXSl7Y2hvc2VuPWZpbGUuZmlsZXNbMF07cGljaygpO319OwpmdW5jdGlvbiBwaWNrKCl7JCgiI2Ryb3B0IikudGV4dENvbnRlbnQ9Y2hvc2VuLm5hbWU7JCgiI2Ryb3BzIikudGV4dENvbnRlbnQ9IlJlYWR5IHRvIGJ1aWxkIjtnby5kaXNhYmxlZD1mYWxzZTt9CnJlc2V0Lm9uY2xpY2s9KCk9PntjaG9zZW49bnVsbDtmaWxlLnZhbHVlPSIiO2dvLmRpc2FibGVkPXRydWU7cmVzZXQuY2xhc3NMaXN0LmFkZCgiaGlkZSIpO3N0YXR1cy50ZXh0Q29udGVudD0iIjtyZXN1bHQuaW5uZXJIVE1MPSIiO2ZlYXRzLmNsYXNzTGlzdC5yZW1vdmUoImhpZGUiKTskKCIjZHJvcHQiKS50ZXh0Q29udGVudD0iRHJvcCB5b3VyIHNwcmVhZHNoZWV0IGhlcmUiOyQoIiNkcm9wcyIpLmlubmVySFRNTD0ib3IgY2xpY2sgdG8gYnJvd3NlICZuYnNwOyZtaWRkb3Q7Jm5ic3A7IC54bHN4IG9yIC54bHMiO307CmNvbnN0IGZtdD1uPT5uLnRvTG9jYWxlU3RyaW5nKCk7CmdvLm9uY2xpY2s9YXN5bmMoKT0+ewogZ28uZGlzYWJsZWQ9dHJ1ZTtyZXN1bHQuaW5uZXJIVE1MPSIiO3N0YXR1cy5pbm5lckhUTUw9JzxzcGFuIGNsYXNzPSJzcGluIj48L3NwYW4+IEJ1aWxkaW5nIHBsYW4nOwogY29uc3QgZmQ9bmV3IEZvcm1EYXRhKCk7ZmQuYXBwZW5kKCJmaWxlIixjaG9zZW4pOwogdHJ5ewogICBjb25zdCByPWF3YWl0IGZldGNoKCIvYXBpL3Byb2Nlc3MiLHttZXRob2Q6IlBPU1QiLGJvZHk6ZmR9KTsKICAgY29uc3Qgaj1hd2FpdCByLmpzb24oKTsKICAgaWYoai5lcnJvcil7c3RhdHVzLnRleHRDb250ZW50PSIiO3Jlc3VsdC5pbm5lckhUTUw9JzxkaXYgY2xhc3M9InJjYXJkIj48ZGl2IGNsYXNzPSJlcnIiPjxiPkNvdWxkbnQgcHJvY2VzcyB0aGlzIGZpbGUuPC9iPjxicj4nK2ouZXJyb3IrJzwvZGl2PjwvZGl2Pic7Z28uZGlzYWJsZWQ9ZmFsc2U7cmV0dXJuO30KICAgc3RhdHVzLnRleHRDb250ZW50PSIiO3Jlc2V0LmNsYXNzTGlzdC5yZW1vdmUoImhpZGUiKTtmZWF0cy5jbGFzc0xpc3QuYWRkKCJoaWRlIik7CiAgIGNvbnN0IHM9ai5zdW1tYXJ5OwogICBjb25zdCBtaXg9T2JqZWN0LmVudHJpZXMocy5lcXVpcG1lbnRfbWl4KS5zb3J0KChhLGIpPT5iWzFdLWFbMV0pLm1hcCgoW2ssdl0pPT4nPHNwYW4gY2xhc3M9InBpbGwyIj48Yj4nK3YrJzwvYj4nK2srJzwvc3Bhbj4nKS5qb2luKCIiKTsKICAgbGV0IGJhbm5lcjsKICAgaWYocy5zdXBlcmxvYWRzLmxlbmd0aCl7Y29uc3QgdT1bLi4ubmV3IFNldChzLnN1cGVybG9hZHMubWFwKHg9PiJSb3cgIit4WzBdKyIgICIreFsxXSkpXTsKICAgICBiYW5uZXI9JzxkaXYgY2xhc3M9ImJhbm5lciB3YXJuIj48aT4mIzk4ODg7PC9pPjxkaXY+PHNwYW4gY2xhc3M9InR0bCI+Jyt1Lmxlbmd0aCsnIHBlcm1pdCAvIHN1cGVybG9hZCBmbGFnJysodS5sZW5ndGg+MT8icyI6IiIpKyc8L3NwYW4+PGJyPicrdS5qb2luKCI8YnI+IikrJzwvZGl2PjwvZGl2Pic7fQogICBlbHNlIGJhbm5lcj0nPGRpdiBjbGFzcz0iYmFubmVyIG9rIj48aT4mIzEwMDAzOzwvaT48ZGl2PjxzcGFuIGNsYXNzPSJ0dGwiPk5vIHN1cGVybG9hZHMuPC9zcGFuPiBOb3RoaW5nIGV4Y2VlZHMgMTYgZnQgd2lkZSwgMTMmIzM5OzYmIzM0OyB0YWxsLCBvciB0aGUgd2VpZ2h0IGNlaWxpbmcuPC9kaXY+PC9kaXY+JzsKICAgY29uc3QgY2VsbD0obixsKT0+JzxkaXYgY2xhc3M9InN0YXQiPjxkaXYgY2xhc3M9Im4iPicrbisnPC9kaXY+PGRpdiBjbGFzcz0ibCI+JytsKyc8L2Rpdj48L2Rpdj4nOwogICByZXN1bHQuaW5uZXJIVE1MPSc8ZGl2IGNsYXNzPSJyY2FyZCI+PGRpdiBjbGFzcz0icmhlYWQiPjxzcGFuIGNsYXNzPSJjaGVjayI+JiMxMDAwMzs8L3NwYW4+IFBsYW4gcmVhZHkgICcrY2hvc2VuLm5hbWUrJzwvZGl2PicrCiAgICAgJzxkaXYgY2xhc3M9InN0YXRzIj4nK2NlbGwocy50cnVja3MsIlRydWNrcyIpK2NlbGwocy5waWVjZXMsIlBpZWNlcyIpK2NlbGwocy5vdmVyX3dpZHRoX3BpZWNlcywiT3Zlci13aWR0aCIpK2NlbGwoTWF0aC5yb3VuZChzLnRvdGFsX3dlaWdodF9sYi8xMDAwKSsiayIsIlRvdGFsIGxiIikrJzwvZGl2PicrCiAgICAgJzxkaXYgY2xhc3M9Im1peCI+JyttaXgrJzwvZGl2PicrYmFubmVyKwogICAgIChzLnNraXBwZWQuY29uc29saWRhdGVkPyc8cCBzdHlsZT0iY29sb3I6dmFyKC0tc3ViKTtmb250LXNpemU6MTNweDttYXJnaW46MCAwIDE2cHgiPicrcy5za2lwcGVkLmNvbnNvbGlkYXRlZCsnIGNvbnNvbGlkYXRlZCBsaW5lKHMpIGxlZnQgYmxhbmsuPC9wPic6IiIpKwogICAgICc8YSBjbGFzcz0iZGwiIGRvd25sb2FkPSInK2ouZmlsZW5hbWUrJyIgaHJlZj0iZGF0YTphcHBsaWNhdGlvbi92bmQub3BlbnhtbGZvcm1hdHMtb2ZmaWNlZG9jdW1lbnQuc3ByZWFkc2hlZXRtbC5zaGVldDtiYXNlNjQsJytqLmZpbGUrJyI+JiMxMTAxNTsgRG93bmxvYWQgdHJ1Y2sgcGxhbiAoLnhsc3gpPC9hPjwvZGl2Pic7CiB9Y2F0Y2goZSl7c3RhdHVzLnRleHRDb250ZW50PSIiO3Jlc3VsdC5pbm5lckhUTUw9JzxkaXYgY2xhc3M9InJjYXJkIj48ZGl2IGNsYXNzPSJlcnIiPjxiPkVycm9yLjwvYj4gJytlKyc8L2Rpdj48L2Rpdj4nO2dvLmRpc2FibGVkPWZhbHNlO30KfTsKPC9zY3JpcHQ+CjwvYm9keT4KPC9odG1sPgo=")

@app.route('/')
def home():
    return Response(_PAGE, mimetype='text/html')

@app.route('/welcome.png')
def welcome_png():
    p = os.path.join(os.path.dirname(__file__), 'welcome.png')
    return send_file(p, mimetype='image/png') if os.path.exists(p) else ('', 404)

@app.route('/api/process', methods=['POST'])
def api_process():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file received.'}), 200
    try:
        out, summary = process(f.read(), f.filename or 'manifest.xlsx')
        base = (f.filename or 'manifest.xlsx').rsplit('.', 1)[0]
        return jsonify({'summary': summary,
                        'filename': base + ' JRCTEST.xlsx',
                        'file': _b64.b64encode(out).decode()})
    except Exception as e:
        return jsonify({'error': str(e)}), 200


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)
